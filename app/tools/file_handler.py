"""ATLAS dosya olusturma modulu.

Excel, CSV ve PDF dosya uretimi saglayan
yeniden kullanilabilir arac sinifi.

Agent'lar bu sinifi kullanarak rapor ve veri dosyalari
olusturabilir. Bagimsiz olarak da kullanilabilir.
"""

import csv
import logging
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from pydantic import BaseModel

logger = logging.getLogger("atlas.tools.file_handler")

# openpyxl opsiyonel import
try:
    import openpyxl as _openpyxl_test  # noqa: F401

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logger.info("openpyxl kurulu degil, Excel olusturma kullanilamaz")

# reportlab opsiyonel import
try:
    import reportlab as _reportlab_test  # noqa: F401

    _REPORTLAB_AVAILABLE = True
except ImportError:
    _REPORTLAB_AVAILABLE = False
    logger.info("reportlab kurulu degil, PDF olusturma kullanilamaz")


class FileResult(BaseModel):
    """Dosya olusturma sonucu.

    Attributes:
        file_path: Olusturulan dosyanin yolu.
        file_type: Dosya tipi (xlsx, pdf, csv).
        file_size: Dosya boyutu (byte).
        success: Olusturma basarili mi.
        error: Hata mesaji (basarisizsa).
    """

    file_path: str = ""
    file_type: str = ""
    file_size: int = 0
    success: bool = True
    error: str = ""


class FileHandler:
    """Dosya olusturma yoneticisi.

    Excel, CSV ve PDF dosya uretimi saglar.
    CSV her zaman kullanilabilir, Excel ve PDF opsiyonel
    kutuphanelere bagimlidir.

    Kullanim:
        handler = FileHandler(output_dir="output/")
        result = handler.create_csv(
            data=[["Ad", "Soyad"], ["Fatih", "Test"]],
            filename="rapor.csv",
        )

    Attributes:
        output_dir: Dosyalarin kaydedilecegi dizin.
    """

    def __init__(self, output_dir: str = "output") -> None:
        """FileHandler'i baslatir.

        Args:
            output_dir: Cikti dizini yolu.
        """
        self.output_dir = output_dir

    def _ensure_output_dir(self) -> None:
        """Cikti dizininin var olmasini saglar."""
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)

    def _full_path(self, filename: str) -> str:
        """Tam dosya yolunu olusturur.

        Args:
            filename: Dosya adi.

        Returns:
            Tam dosya yolu.
        """
        return str(Path(self.output_dir) / filename)

    def _get_file_size(self, path: str) -> int:
        """Dosya boyutunu dondurur.

        Args:
            path: Dosya yolu.

        Returns:
            Dosya boyutu (byte).
        """
        try:
            return os.path.getsize(path)
        except OSError:
            return 0

    def create_csv(
        self,
        data: list[list[Any]],
        filename: str = "export.csv",
        headers: list[str] | None = None,
    ) -> FileResult:
        """CSV dosyasi olusturur.

        Her zaman kullanilabilir (stdlib csv modulu).

        Args:
            data: Satir listesi. Her satir deger listesi.
            filename: Dosya adi.
            headers: Sutun basliklari (opsiyonel, data'nin ilk
                satiri olarak da verilebilir).

        Returns:
            Dosya olusturma sonucu.
        """
        self._ensure_output_dir()
        file_path = self._full_path(filename)

        try:
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                if headers:
                    writer.writerow(headers)
                writer.writerows(data)

            file_size = self._get_file_size(file_path)
            logger.info("CSV olusturuldu: %s (%d byte)", file_path, file_size)

            return FileResult(
                file_path=file_path,
                file_type="csv",
                file_size=file_size,
                success=True,
            )
        except Exception as exc:
            logger.error("CSV olusturma hatasi: %s", exc)
            return FileResult(
                file_path=file_path,
                file_type="csv",
                success=False,
                error=str(exc),
            )

    def create_excel(
        self,
        data: list[list[Any]],
        filename: str = "export.xlsx",
        sheet_name: str = "Sheet1",
        headers: list[str] | None = None,
        column_widths: list[int] | None = None,
    ) -> FileResult:
        """Excel (xlsx) dosyasi olusturur.

        openpyxl kutuphanesi gereklidir. Header satiri
        bold ve renkli arka plan ile formatlanir.

        Args:
            data: Satir listesi. Her satir deger listesi.
            filename: Dosya adi.
            sheet_name: Sayfa adi.
            headers: Sutun basliklari (opsiyonel).
            column_widths: Sutun genislikleri (opsiyonel).

        Returns:
            Dosya olusturma sonucu.

        Raises:
            RuntimeError: openpyxl kurulu degilse.
        """
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl kurulu degil. "
                "Kurmak icin: pip install openpyxl"
            )

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        self._ensure_output_dir()
        file_path = self._full_path(filename)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Header formatlama
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid",
            )

            start_row = 1
            if headers:
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                start_row = 2

            # Veri satirlari
            for row_idx, row_data in enumerate(data, start_row):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Sutun genislikleri
            if column_widths:
                for col_idx, width in enumerate(column_widths, 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = width
            else:
                # Otomatik genislik
                self._auto_column_width(ws)

            wb.save(file_path)
            file_size = self._get_file_size(file_path)
            logger.info("Excel olusturuldu: %s (%d byte)", file_path, file_size)

            return FileResult(
                file_path=file_path,
                file_type="xlsx",
                file_size=file_size,
                success=True,
            )
        except Exception as exc:
            logger.error("Excel olusturma hatasi: %s", exc)
            return FileResult(
                file_path=file_path,
                file_type="xlsx",
                success=False,
                error=str(exc),
            )

    def create_pdf(
        self,
        title: str,
        content_lines: list[str],
        filename: str = "report.pdf",
        author: str = "ATLAS",
    ) -> FileResult:
        """PDF rapor dosyasi olusturur.

        reportlab kutuphanesi gereklidir. Baslik, tarih,
        yazar ve satir satir icerik ile PDF olusturur.

        Args:
            title: Rapor basligi.
            content_lines: Icerik satirlari.
            filename: Dosya adi.
            author: Rapor yazari.

        Returns:
            Dosya olusturma sonucu.

        Raises:
            RuntimeError: reportlab kurulu degilse.
        """
        if not _REPORTLAB_AVAILABLE:
            raise RuntimeError(
                "reportlab kurulu degil. "
                "Kurmak icin: pip install reportlab"
            )

        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.pdfgen import canvas

        self._ensure_output_dir()
        file_path = self._full_path(filename)

        try:
            c = canvas.Canvas(file_path, pagesize=A4)
            width, height = A4

            # Baslik
            c.setFont("Helvetica-Bold", 16)
            c.drawString(2 * cm, height - 2 * cm, title)

            # Tarih ve yazar
            c.setFont("Helvetica", 9)
            now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
            c.drawString(2 * cm, height - 2.8 * cm, f"Tarih: {now} | Yazar: {author}")

            # Ayirici cizgi
            c.setLineWidth(0.5)
            c.line(2 * cm, height - 3 * cm, width - 2 * cm, height - 3 * cm)

            # Icerik
            c.setFont("Helvetica", 10)
            y = height - 3.8 * cm
            line_height = 14

            for line in content_lines:
                if y < 2 * cm:
                    # Sayfa numarasi
                    c.setFont("Helvetica", 8)
                    c.drawCentredString(
                        width / 2, 1 * cm,
                        f"Sayfa {c.getPageNumber()}",
                    )
                    c.showPage()
                    c.setFont("Helvetica", 10)
                    y = height - 2 * cm

                c.drawString(2 * cm, y, line)
                y -= line_height

            # Son sayfa numarasi
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2, 1 * cm, f"Sayfa {c.getPageNumber()}")

            c.save()
            file_size = self._get_file_size(file_path)
            logger.info("PDF olusturuldu: %s (%d byte)", file_path, file_size)

            return FileResult(
                file_path=file_path,
                file_type="pdf",
                file_size=file_size,
                success=True,
            )
        except Exception as exc:
            logger.error("PDF olusturma hatasi: %s", exc)
            return FileResult(
                file_path=file_path,
                file_type="pdf",
                success=False,
                error=str(exc),
            )

    def create_report(
        self,
        title: str,
        sections: list[dict[str, Any]],
        filename: str | None = None,
        file_type: str = "xlsx",
    ) -> FileResult:
        """Yapilandirilmis rapor dosyasi olusturur.

        Birden fazla bolumlu raporlar icin ust seviye metod.
        Her bolum baslik ve veri icerir.

        Args:
            title: Rapor basligi.
            sections: Rapor bolumleri.
                Her bolum: {"title": str, "headers": list, "data": list[list]}
            filename: Dosya adi (None ise otomatik).
            file_type: Dosya tipi ("xlsx" veya "pdf").

        Returns:
            Dosya olusturma sonucu.
        """
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        auto_filename = filename or f"rapor_{timestamp}.{file_type}"

        if file_type == "pdf":
            # PDF rapor: tum bolumleri metin olarak birlestir
            lines: list[str] = []
            for section in sections:
                lines.append("")
                lines.append(f"--- {section.get('title', 'Bolum')} ---")
                lines.append("")

                section_headers = section.get("headers", [])
                if section_headers:
                    lines.append("  ".join(str(h) for h in section_headers))
                    lines.append("-" * 60)

                for row in section.get("data", []):
                    lines.append("  ".join(str(v) for v in row))

            return self.create_pdf(
                title=title,
                content_lines=lines,
                filename=auto_filename,
            )

        # Excel rapor: her bolum ayri sayfa
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "openpyxl kurulu degil. "
                "Kurmak icin: pip install openpyxl"
            )

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        self._ensure_output_dir()
        file_path = self._full_path(auto_filename)

        try:
            wb = openpyxl.Workbook()
            # Varsayilan sayfa sil
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid",
            )

            for section in sections:
                section_title = section.get("title", "Sheet")[:31]
                ws = wb.create_sheet(title=section_title)

                section_headers = section.get("headers", [])
                start_row = 1

                if section_headers:
                    for col_idx, header in enumerate(section_headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    start_row = 2

                for row_idx, row_data in enumerate(section.get("data", []), start_row):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                self._auto_column_width(ws)

            wb.save(file_path)
            file_size = self._get_file_size(file_path)
            logger.info("Rapor olusturuldu: %s (%d byte)", file_path, file_size)

            return FileResult(
                file_path=file_path,
                file_type=file_type,
                file_size=file_size,
                success=True,
            )
        except Exception as exc:
            logger.error("Rapor olusturma hatasi: %s", exc)
            return FileResult(
                file_path=file_path,
                file_type=file_type,
                success=False,
                error=str(exc),
            )

    # === Yardimci metodlar ===

    @staticmethod
    def _auto_column_width(ws: Any) -> None:
        """Sayfa sutun genisliklerini otomatik ayarlar.

        Args:
            ws: openpyxl Worksheet nesnesi.
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column_cells:
                if column_letter is None:
                    column_letter = cell.column_letter
                try:
                    cell_length = len(str(cell.value or ""))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            if column_letter:
                adjusted = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted
